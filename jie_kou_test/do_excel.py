import unittest
from openpyxl import load_workbook

class DoExcel:

    def __init__(self,excle_name,sheet_name):
        self.excle_name = excle_name
        self.sheet_name = sheet_name

    def do_excel(self):
        wb = load_workbook(self.excle_name)
        sheet = wb[self.sheet_name]

        data_dict = []
        for i in range(2, sheet.max_row + 1):
            data_list = {}

            data_list["method"] = sheet.cell(i, 1).value
            data_list["url"] = sheet.cell(i, 2).value
            data_list["data"] = eval(sheet.cell(i, 3).value)

            # print(data_list)
            data_dict.append(data_list)

        # print(data_dict)
        return data_dict

if __name__ == '__main__':
    data = DoExcel("python.xlsx", "test").do_excel()
    print(data)